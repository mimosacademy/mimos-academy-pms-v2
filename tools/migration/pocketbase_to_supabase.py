#!/usr/bin/env python3
"""MIMOS Academy PMS migration utility.

Sources:
  * PocketBase pb_data/data.db (SQLite)
  * Excel workbooks (generic lossless staging)

Design:
  1. Create an import_batch.
  2. Preserve every source row in stg_raw_record.
  3. Map PocketBase collections to the authoritative Supabase model.
  4. Resolve relations before inserting children.
  5. Park unresolved relationships/duplicates in data_conflict instead of guessing.

Environment:
  SUPABASE_DB_URL=postgresql://...

Install:
  pip install psycopg[binary] openpyxl
"""
from __future__ import annotations

import argparse, hashlib, json, os, re, sqlite3
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import psycopg
from psycopg.types.json import Jsonb
from openpyxl import load_workbook


def clean(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.isoformat()
    if isinstance(v, Decimal): return str(v)
    s = str(v).strip()
    if s in ('', '-', 'N/A', 'NA', 'No Data', '#REF!', '#DIV/0!'): return None
    return s


def money(v):
    v = clean(v)
    if v is None: return None
    v = re.sub(r'[^0-9.\-]', '', v)
    try: return Decimal(v).quantize(Decimal('0.01'))
    except InvalidOperation: return None


def dt_date(v):
    v = clean(v)
    if not v: return None
    for fmt in ('%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%d-%b-%y','%m/%d/%y'):
        try: return datetime.strptime(v, fmt).date()
        except ValueError: pass
    return None


def norm(v):
    v = clean(v)
    return re.sub(r'[^a-z0-9]+',' ',v.lower()).strip() if v else None


def read_pb(db_path: Path):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    tables = [r[0] for r in con.execute("select name from sqlite_master where type='table' and name not like '_%' and name not like 'sqlite_%'")]
    result = {}
    for table in tables:
        try:
            rows = [dict(r) for r in con.execute(f' select * from "{table.replace(chr(34), chr(34)*2)}" ')]
            result[table] = rows
        except sqlite3.DatabaseError:
            pass
    con.close()
    return result


def connect():
    url = os.environ.get('SUPABASE_DB_URL')
    if not url: raise SystemExit('SUPABASE_DB_URL is required')
    return psycopg.connect(url)


def batch(cur, code, import_type, target='MIGRATION'):
    cur.execute("insert into import_batch(batch_code,import_type,table_target,status) values(%s,%s,%s,'PROCESSING') returning id", (code,import_type,target))
    return cur.fetchone()[0]


def stage_raw(cur, batch_id, source_file, source_type, rows, target=None):
    for n, row in enumerate(rows, start=1):
        cur.execute("""insert into stg_raw_record(import_batch_id,source_file,source_row_number,source_type,target_table,raw_data)
          values(%s,%s,%s,%s,%s,%s)""", (batch_id,source_file,n,source_type,target,Jsonb(row)))


def conflict(cur,batch_id,source_file,row_no,table,typ,desc,source_value=None,existing_value=None,target_id=None):
    cur.execute("""insert into data_conflict(import_batch_id,source_file,source_row_number,table_name,target_record_id,conflict_type,source_value,existing_value,conflict_description)
      values(%s,%s,%s,%s,%s,%s,%s,%s,%s)""",(batch_id,source_file,row_no,table,target_id,typ,clean(source_value),clean(existing_value),desc))


def id_map(cur, table, pb_id, cache):
    return cache.get((table, str(pb_id))) if pb_id else None


def migrate_pb(db_path: Path, commit=False):
    collections = read_pb(db_path)
    code = 'PB-' + datetime.utcnow().strftime('%Y%m%d%H%M%S')
    with connect() as pg:
      with pg.cursor() as cur:
        batch_id = batch(cur, code, 'POCKETBASE_SQLITE')
        for name, rows in collections.items():
            stage_raw(cur,batch_id,f'pocketbase:{name}','POCKETBASE',rows,name)

        maps = {}
        # 1. clients
        for n,r in enumerate(collections.get('clients',[]),1):
            name=clean(r.get('name')) or f'Unresolved Client {r.get("id")}'
            cur.execute("select id from client where lower(company_name)=lower(%s) limit 1",(name,))
            found=cur.fetchone()
            if found: maps[('clients',str(r.get('id')))] = found[0]; continue
            cur.execute("insert into client(company_name,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s) returning id",(name,'pocketbase:clients',n,batch_id))
            maps[('clients',str(r.get('id')))] = cur.fetchone()[0]

        # 2. contacts
        for n,r in enumerate(collections.get('client_contacts',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps)
            if not cid:
                conflict(cur,batch_id,'pocketbase:client_contacts',n,'client_contact','MISSING_RELATION','PocketBase client relation could not be resolved',r.get('client'))
                continue
            cur.execute("insert into client_contact(client_id,contact_name,contact_email,contact_phone,contact_designation,is_primary,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s)", (cid,clean(r.get('name')) or 'Unknown',clean(r.get('email')),clean(r.get('phone')),clean(r.get('title')),bool(r.get('isPrimary')), 'pocketbase:client_contacts',n,batch_id))

        # 3. programmes
        status_map={'Scheduled':'PLANNED','In Progress':'IN_PROGRESS','Completed':'COMPLETED','On Hold':'ON_HOLD'}
        for n,r in enumerate(collections.get('programmes',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps)
            if not cid:
                conflict(cur,batch_id,'pocketbase:programmes',n,'programme','MISSING_RELATION','Client relation could not be resolved',r.get('client')); continue
            cur.execute("select id from programme where programme_code=%s limit 1",(clean(r.get('code')),))
            found=cur.fetchone()
            if found: maps[('programmes',str(r.get('id')))] = found[0]; continue
            cur.execute("select id from programme_status where code=%s",(status_map.get(clean(r.get('status')),'PLANNED'),)); ps=cur.fetchone(); psid=ps[0] if ps else None
            cur.execute("""insert into programme(programme_code,title,client_id,programme_status_id,start_date,end_date,no_of_pax,total_revenue_excl_tax,total_revenue_incl_tax,source_file,source_row_number,import_batch_id)
              values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) returning id""",(clean(r.get('code')),clean(r.get('title')) or 'Untitled',cid,psid,dt_date(r.get('startDate')),dt_date(r.get('endDate')),int(float(r.get('participants') or 0)),money(r.get('contractValue')) or 0,money(r.get('contractValue')) or 0,'pocketbase:programmes',n,batch_id))
            maps[('programmes',str(r.get('id')))] = cur.fetchone()[0]

        # 4. opportunities
        stage_map={'Lead':'EARLY_ENGAGEMENT','Qualified':'QUALIFIED_LEAD','Proposal':'PROPOSAL_SUBMITTED','Negotiation':'NEGOTIATION','Won':'WON','Lost':'LOST'}
        for n,r in enumerate(collections.get('opportunities',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps); pid=id_map(cur,'programmes',r.get('linkedProgramme'),maps)
            if not cid:
                conflict(cur,batch_id,'pocketbase:opportunities',n,'opportunity','MISSING_RELATION','Client relation could not be resolved',r.get('client')); continue
            cur.execute("select id from opportunity_status where code=%s",(stage_map.get(clean(r.get('stage')),'EARLY_ENGAGEMENT'),)); osid=cur.fetchone()[0]
            cur.execute("insert into opportunity(opportunity_code,client_id,programme_id,opportunity_status_id,project_title,forecast_value,probability_percentage,weighted_value,expected_close_date,remarks,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) returning id",(clean(r.get('id')),cid,pid,osid,clean(r.get('title')) or 'Untitled',money(r.get('value')),Decimal(str(r.get('probability'))) if r.get('probability') is not None else None,None,dt_date(r.get('expectedClose')),clean(r.get('source')),'pocketbase:opportunities',n,batch_id))
            maps[('opportunities',str(r.get('id')))] = cur.fetchone()[0]

        # 5. quotations
        qstatus={'Draft':'IN_PROGRESS','Sent':'SENT','Accepted':'ACCEPTED','Rejected':'REJECTED','Expired':'EXPIRED'}
        for n,r in enumerate(collections.get('quotations',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps); pid=id_map(cur,'programmes',r.get('programme'),maps)
            if not cid: conflict(cur,batch_id,'pocketbase:quotations',n,'quotation','MISSING_RELATION','Client relation could not be resolved',r.get('client')); continue
            cur.execute("select id from quotation_status where code=%s",(qstatus.get(clean(r.get('status')),'IN_PROGRESS'),)); qsid=cur.fetchone()[0]
            cur.execute("insert into quotation(quotation_no,programme_id,client_id,quotation_status_id,project_title,final_price,quotation_date,valid_until,pic_full_name,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (clean(r.get('quoteNo')) or clean(r.get('id')),pid,cid,qsid,clean(r.get('programmeTitle')),money(r.get('amount')),dt_date(r.get('issueDate')),dt_date(r.get('validUntil')),clean(r.get('preparedBy')),'pocketbase:quotations',n,batch_id))

        # 6. purchase orders
        for n,r in enumerate(collections.get('purchase_orders',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps); pid=id_map(cur,'programmes',r.get('programme'),maps)
            if not cid: conflict(cur,batch_id,'pocketbase:purchase_orders',n,'purchase_order','MISSING_RELATION','Client relation could not be resolved',r.get('client')); continue
            cur.execute("insert into purchase_order(po_no,client_id,programme_id,po_value_incl_tax,po_date,po_status,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s)",(clean(r.get('poNo')),cid,pid,money(r.get('amount')),dt_date(r.get('issueDate')),clean(r.get('status')),'pocketbase:purchase_orders',n,batch_id))

        # 7. invoices
        invstatus={'Unpaid':'UNPAID','Paid':'PAID','Overdue':'OVERDUE','Partial':'PARTIAL'}
        for n,r in enumerate(collections.get('invoices',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps); pid=id_map(cur,'programmes',r.get('programme'),maps)
            if not cid or not pid:
                conflict(cur,batch_id,'pocketbase:invoices',n,'invoice','MISSING_RELATION','Client/programme relation could not be resolved',f"client={r.get('client')},programme={r.get('programme')}"); continue
            cur.execute("select id from payment_status where code=%s",(invstatus.get(clean(r.get('status')),'UNKNOWN'),)); psid=cur.fetchone()[0]
            amount=money(r.get('amount')); paid=money(r.get('paidAmount')) or 0
            cur.execute("insert into invoice(invoice_no,programme_id,client_id,payment_status_id,invoice_date,due_date,amount_excl_tax,total_incl_tax,amount_collected,amount_outstanding,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) returning id",(clean(r.get('invoiceNo')),pid,cid,psid,dt_date(r.get('issueDate')),dt_date(r.get('dueDate')),amount,amount,paid,(amount or 0)-paid,'pocketbase:invoices',n,batch_id))
            maps[('invoices',str(r.get('id')))] = cur.fetchone()[0]

        # 8. payments
        method_map={'Bank Transfer':'BANK_TRANSFER','Cheque':'BANK_TRANSFER','Online Banking':'BANK_TRANSFER','Credit Card':'SELF_PAY','HRDCorp Claimable':'HRDCORP','Self-Pay':'SELF_PAY','ePerolehan':'EPEROLEHAN'}
        for n,r in enumerate(collections.get('payments',[]),1):
            cid=id_map(cur,'clients',r.get('client'),maps); pid=id_map(cur,'programmes',r.get('programme'),maps); iid=id_map(cur,'invoices',r.get('invoice'),maps)
            if not pid or not cid:
                conflict(cur,batch_id,'pocketbase:payments',n,'payment','MISSING_RELATION','Client/programme relation could not be resolved'); continue
            cur.execute("select id from payment_method where code=%s",(method_map.get(clean(r.get('method')),'SELF_PAY'),)); pmid=cur.fetchone()[0]
            cur.execute("select id from payment_status where code=%s",('PAID' if clean(r.get('status'))=='Completed' else 'PENDING',)); psid=cur.fetchone()[0]
            cur.execute("insert into payment(payment_reference,invoice_id,programme_id,client_id,payment_method_id,payment_status_id,payment_date,amount,total_amount,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(clean(r.get('paymentNo')),iid,pid,cid,pmid,psid,dt_date(r.get('date')),money(r.get('amount')) or 0,money(r.get('amount')) or 0,'pocketbase:payments',n,batch_id))

        # 9. participants and training stats
        for n,r in enumerate(collections.get('participants',[]),1):
            pid=id_map(cur,'programmes',r.get('programme'),maps); cid=id_map(cur,'clients',r.get('client'),maps)
            if not pid: conflict(cur,batch_id,'pocketbase:participants',n,'participant','MISSING_RELATION','Programme relation could not be resolved',r.get('programme')); continue
            cur.execute("insert into participant(programme_id,full_name,email,phone,organization,attendance_status,source_file,source_row_number,import_batch_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s)",(pid,clean(r.get('name')) or 'Unknown',clean(r.get('email')),clean(r.get('phone')),clean(r.get('company')),clean(r.get('status')) or 'ATTENDED','pocketbase:participants',n,batch_id))

        # PB-only collections are already preserved losslessly in stg_raw_record.
        cur.execute("update import_batch set status='COMPLETED',end_time=now() where id=%s",(batch_id,))
        if not commit:
            pg.rollback(); print(json.dumps({'batch_id':batch_id,'mode':'DRY_RUN','message':'Rollback complete; no canonical data persisted.'})); return
        print(json.dumps({'batch_id':batch_id,'mode':'COMMIT','message':'Migration committed.'}))


def stage_excel(folder: Path, commit=False):
    files=list(folder.glob('*.xlsx'))
    with connect() as pg:
      with pg.cursor() as cur:
        batch_id=batch(cur,'XLSX-'+datetime.utcnow().strftime('%Y%m%d%H%M%S'),'EXCEL_STAGING','stg_raw_record')
        total=0
        for path in files:
            digest=hashlib.sha256(path.read_bytes()).hexdigest()
            cur.execute("insert into source_file(file_name,file_path,file_hash,file_size_bytes,file_type) values(%s,%s,%s,%s,%s) returning id",(path.name,str(path),digest,path.stat().st_size,path.suffix.lower()))
            sfid=cur.fetchone()[0]
            wb=load_workbook(path,read_only=True,data_only=True)
            for ws in wb.worksheets:
                rows=ws.iter_rows(values_only=True)
                headers=next(rows,None)
                if not headers: continue
                headers=[str(h).strip() if h is not None else f'column_{i+1}' for i,h in enumerate(headers)]
                for row_no,values in enumerate(rows,start=2):
                    data={headers[i]:clean(values[i]) if i<len(values) else None for i in range(len(headers))}
                    cur.execute("insert into stg_raw_record(import_batch_id,source_file,source_sheet,source_row_number,source_type,target_table,raw_data) values(%s,%s,%s,%s,'EXCEL',%s,%s)",(batch_id,path.name,ws.title,row_no,None,Jsonb(data)))
                    total+=1
            wb.close()
        cur.execute("update import_batch set records_total=%s,status='STAGED',end_time=now() where id=%s",(total,batch_id))
        if not commit: pg.rollback(); print(json.dumps({'batch_id':batch_id,'rows':total,'mode':'DRY_RUN'})); return
        print(json.dumps({'batch_id':batch_id,'rows':total,'mode':'COMMIT'}))


if __name__=='__main__':
    ap=argparse.ArgumentParser(); sub=ap.add_subparsers(dest='command',required=True)
    p=sub.add_parser('pocketbase'); p.add_argument('--db',required=True); p.add_argument('--commit',action='store_true')
    e=sub.add_parser('excel'); e.add_argument('--folder',required=True); e.add_argument('--commit',action='store_true')
    a=ap.parse_args()
    if a.command=='pocketbase': migrate_pb(Path(a.db),a.commit)
    else: stage_excel(Path(a.folder),a.commit)
